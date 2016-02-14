from openpyxl import load_workbook


def passportDetail(inputId):
    wb = load_workbook('Book1.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    c = sheet['A2'].value

    passportDetails = ""
    print inputId
    for i in range(2, 836):
        empPhone = sheet.cell(row=i, column=3).value
        if empPhone == int(inputId):
            passportDetails += sheet.cell(row=i, column=2).value
            passportDetails += " Your passport application number is "
            passportDetails += sheet.cell(row=i, column=13).value
            passportDetails += ". Your application status is '"
            passportDetails += sheet.cell(row=i, column=11).value
            passportDetails += " '."
            return passportDetails
    return passportDetails
