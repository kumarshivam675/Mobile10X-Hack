from openpyxl import load_workbook

def cabDetail(inputId):
	wb = load_workbook('Book1.xlsx')
	sheet = wb.get_sheet_by_name('Sheet1')
	c = sheet['A2'].value

	cabDetails = ""
	print inputId
	for i in range(2,836):
		empPhone = sheet.cell(row = i,column = 3).value
		if empPhone == int(inputId):
			gender = sheet.cell(row = i,column = 5).value
			print gender
			cabDetails += sheet.cell(row = i,column = 2).value
			cabDetails += " your Cab Number is "
			cabDetails += sheet.cell(row = i,column = 11).value
			cabDetails += ". "
			if gender == 'M':
				phone = sheet.cell(row = i,column = 13).value
				cabDetails += "Your Drivers phone number is "
				cabDetails += str(phone)
			return cabDetails
	return cabDetails

